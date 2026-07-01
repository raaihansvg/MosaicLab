import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .constants import GRID_SIZE, WARNA_DEFINISI, WARNA_KEYS, BACKGROUND_COLORS, EXCEL_FILLS

def generate_participant_base() -> pd.DataFrame:
    rows, cols = [], []
    for r in range(1, GRID_SIZE + 1):
        for c in range(1, GRID_SIZE + 1):
            rows.append(r)
            cols.append(c)
    koordinat = [f"({r},{c})" for r, c in zip(rows, cols)]
    return pd.DataFrame({"Koordinat": koordinat, "Baris": rows, "Kolom": cols})


def determine_posisi(color_name: str, bg_color: str = None) -> tuple:
    effective_bg = BACKGROUND_COLORS.copy()
    if bg_color:
        effective_bg.add(bg_color)
    if color_name in effective_bg:
        return ("Duduk", "-", "tidak turun")
    else:
        return ("Berdiri", "Mulai", "Turun")


def build_participant_dataframe(formations_data: list, bg_color: str = None) -> pd.DataFrame:
    df = generate_participant_base()
    for i, fdata in enumerate(formations_data, start=1):
        label  = f"F{i}"
        cn_arr = fdata["color_names"].reshape(-1)
        warna_col, posisi_col, naik_col, turun_col = [], [], [], []
        for cn in cn_arr:
            posisi, naik, turun = determine_posisi(cn, bg_color=bg_color)
            warna_col.append(WARNA_DEFINISI[cn]["label"])
            posisi_col.append(posisi)
            naik_col.append(naik)
            turun_col.append(turun)
        df[f"{label}_Warna"]        = warna_col
        df[f"{label}_Posisi"]       = posisi_col
        df[f"{label}_KomandoNaik"]  = naik_col
        df[f"{label}_KomandoTurun"] = turun_col
    return df


def _make_thin_border():
    thin = Side(style="thin", color="FFCCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def export_to_excel(df: pd.DataFrame, formations_data: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Peserta"

    hdr_font_main = Font(bold=True, color="FFFFFFFF", size=10, name="Calibri")
    hdr_font_sub  = Font(bold=True, color="FF1a1a2e", size=9,  name="Calibri")
    hdr_fill_main = PatternFill("solid", fgColor="FF1a1a2e")
    hdr_fill_sub  = PatternFill("solid", fgColor="FFe0e0f0")
    center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border   = _make_thin_border()

    n_formations = sum(1 for col in df.columns if col.endswith("_Warna"))

    ws.merge_cells("A1:A2"); ws["A1"] = "Koordinat"
    ws.merge_cells("B1:B2"); ws["B1"] = "Baris"
    ws.merge_cells("C1:C2"); ws["C1"] = "Kolom"

    fixed_cols = 3
    for idx in range(1, n_formations + 1):
        start = fixed_cols + (idx - 1) * 4 + 1
        end   = start + 3
        ws.merge_cells(f"{get_column_letter(start)}1:{get_column_letter(end)}1")
        cell = ws[f"{get_column_letter(start)}1"]
        cell.value = f"Formasi {idx}"

    for idx in range(1, n_formations + 1):
        start = fixed_cols + (idx - 1) * 4 + 1
        sub_labels = ["Warna", "Posisi", "Komando Naik", "Komando Turun"]
        for j, sub in enumerate(sub_labels):
            ws[f"{get_column_letter(start + j)}2"] = sub

    for row in ws.iter_rows(min_row=1, max_row=2,
                             min_col=1, max_col=fixed_cols + n_formations * 4):
        for cell in row:
            if cell.row == 1:
                cell.font = hdr_font_main
                cell.fill = hdr_fill_main
            else:
                cell.font = hdr_font_sub
                cell.fill = hdr_fill_sub
            cell.alignment = center_align
            cell.border    = thin_border

    ws.freeze_panes = "D3"

    row_fill_alt  = PatternFill("solid", fgColor="FFF5F5FA")
    data_font     = Font(size=9, name="Calibri")
    coord_font    = Font(size=9, name="Courier New", bold=True)

    label_to_key = {v["label"]: k for k, v in WARNA_DEFINISI.items()}
    color_fills = {k: PatternFill("solid", fgColor=EXCEL_FILLS[k]) for k in WARNA_KEYS}
    color_fonts = {k: Font(size=9, name="Calibri", color=("FFFFFFFF" if k in ("hitam", "merah", "biru") else "FF000000"), bold=True) for k in WARNA_KEYS}
    posisi_fill_berdiri = PatternFill("solid", fgColor="FFd4edda")
    posisi_fill_duduk = PatternFill("solid", fgColor="FFf8d7da")

    columns = df.columns.tolist()
    col_is_warna = [c.endswith("_Warna") for c in columns]
    col_is_posisi = [c.endswith("_Posisi") for c in columns]

    for r_idx, row_data in enumerate(df.itertuples(index=False), start=3):
        is_alt = (r_idx % 2 == 0)
        for c_idx, val in enumerate(row_data):
            cell = ws.cell(row=r_idx, column=c_idx + 1)
            cell.value  = val
            cell.border = thin_border
            col_name = columns[c_idx]
            
            if col_name == "Koordinat":
                cell.font      = coord_font
                cell.alignment = center_align
            elif col_name in ("Baris", "Kolom"):
                cell.font      = data_font
                cell.alignment = center_align
            elif col_is_warna[c_idx]:
                warna_key = label_to_key.get(val)
                if warna_key:
                    cell.fill = color_fills[warna_key]
                    cell.font = color_fonts[warna_key]
                cell.alignment = center_align
            elif col_is_posisi[c_idx]:
                cell.font      = data_font
                cell.alignment = center_align
                cell.fill = posisi_fill_berdiri if val == "Berdiri" else posisi_fill_duduk
            else:
                cell.font      = data_font
                cell.alignment = center_align
                if is_alt and not col_is_warna[c_idx]:
                    cell.fill = row_fill_alt

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 7
    ws.column_dimensions["C"].width = 7
    for idx in range(1, n_formations + 1):
        start = fixed_cols + (idx - 1) * 4 + 1
        ws.column_dimensions[get_column_letter(start)].width     = 10
        ws.column_dimensions[get_column_letter(start + 1)].width = 9
        ws.column_dimensions[get_column_letter(start + 2)].width = 13
        ws.column_dimensions[get_column_letter(start + 3)].width = 14

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    for r in range(3, len(df) + 3):
        ws.row_dimensions[r].height = 14

    # ── Sheet 2: Ringkasan Distribusi ─────────────────────────────────────
    ws2 = wb.create_sheet("Ringkasan Distribusi")

    ws2["A1"] = "Ringkasan Distribusi Warna per Formasi"
    ws2["A1"].font      = Font(bold=True, size=13, name="Calibri", color="FF1a1a2e")
    ws2["A1"].alignment = Alignment(horizontal="center")
    last_col_letter = get_column_letter(n_formations + 2)
    ws2.merge_cells(f"A1:{last_col_letter}1")

    sub_row = 2
    ws2.cell(sub_row, 1, "Warna")
    for i in range(1, n_formations + 1):
        ws2.cell(sub_row, i + 1, f"Formasi {i}")
    ws2.cell(sub_row, n_formations + 2, "Total")

    for c in range(1, n_formations + 3):
        cell = ws2.cell(sub_row, c)
        cell.font      = hdr_font_main
        cell.fill      = hdr_fill_main
        cell.alignment = center_align
        cell.border    = thin_border

    row_fill_alt2 = PatternFill("solid", fgColor="FFF5F5FA")
    for wi, wkey in enumerate(WARNA_KEYS):
        r = sub_row + 1 + wi
        ws2.cell(r, 1, WARNA_DEFINISI[wkey]["label"])
        ws2.cell(r, 1).fill = PatternFill("solid", fgColor=EXCEL_FILLS[wkey])
        txt_c = ("FFFFFFFF" if wkey in ("hitam", "merah", "biru") else "FF000000")
        ws2.cell(r, 1).font      = Font(bold=True, size=9, color=txt_c, name="Calibri")
        ws2.cell(r, 1).alignment = center_align
        ws2.cell(r, 1).border    = thin_border

        row_total = 0
        for fi, fdata in enumerate(formations_data):
            cnt = fdata["distribution"].get(wkey, 0)
            row_total += cnt
            cell = ws2.cell(r, fi + 2, cnt)
            cell.font      = data_font
            cell.alignment = center_align
            cell.border    = thin_border
            if wi % 2 == 1:
                cell.fill = row_fill_alt2

        cell_total = ws2.cell(r, n_formations + 2, row_total)
        cell_total.font      = Font(bold=True, size=9, name="Calibri")
        cell_total.alignment = center_align
        cell_total.border    = thin_border

    tot_row = sub_row + len(WARNA_KEYS) + 1
    ws2.cell(tot_row, 1, "TOTAL")
    ws2.cell(tot_row, 1).font      = Font(bold=True, size=9)
    ws2.cell(tot_row, 1).alignment = center_align
    ws2.cell(tot_row, 1).border    = thin_border
    grand = 0
    for fi, fdata in enumerate(formations_data):
        s = sum(fdata["distribution"].values())
        grand += s
        c = ws2.cell(tot_row, fi + 2, s)
        c.font      = Font(bold=True, size=9)
        c.alignment = center_align
        c.border    = thin_border
    ws2.cell(tot_row, n_formations + 2, grand).font = Font(bold=True, size=9)
    ws2.cell(tot_row, n_formations + 2).alignment = center_align
    ws2.cell(tot_row, n_formations + 2).border = thin_border

    ws2.column_dimensions["A"].width = 12
    for i in range(2, n_formations + 3):
        ws2.column_dimensions[get_column_letter(i)].width = 11

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
