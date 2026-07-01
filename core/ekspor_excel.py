import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .konstanta import UKURAN_GRID, DEFINISI_WARNA, DAFTAR_WARNA, WARNA_LATAR, WARNA_EXCEL


def buat_dataframe_dasar() -> pd.DataFrame:
    daftarBaris, daftarKolom = [], []
    for b in range(1, UKURAN_GRID + 1):
        for k in range(1, UKURAN_GRID + 1):
            daftarBaris.append(b)
            daftarKolom.append(k)
    koordinat = [f"({b},{k})" for b, k in zip(daftarBaris, daftarKolom)]
    return pd.DataFrame({"Koordinat": koordinat, "Baris": daftarBaris, "Kolom": daftarKolom})


def tentukan_posisi(namaWarna: str, warnaBg: str = None) -> tuple:
    latarEfektif = WARNA_LATAR.copy()
    if warnaBg:
        latarEfektif.add(warnaBg)
    if namaWarna in latarEfektif:
        return ("Duduk", "-", "tidak turun")
    else:
        return ("Berdiri", "Mulai", "Turun")


def bangun_dataframe_peserta(daftarFormasi: list, warnaBg: str = None) -> pd.DataFrame:
    df = buat_dataframe_dasar()
    for i, dataFormasi in enumerate(daftarFormasi, start=1):
        labelFormasi = f"F{i}"
        arrayNamaWarna = dataFormasi["nama_warna"].reshape(-1)
        kolomWarna, kolomPosisi, kolomNaik, kolomTurun = [], [], [], []
        for nama in arrayNamaWarna:
            posisi, naik, turun = tentukan_posisi(nama, warnaBg=warnaBg)
            kolomWarna.append(DEFINISI_WARNA[nama]["label"])
            kolomPosisi.append(posisi)
            kolomNaik.append(naik)
            kolomTurun.append(turun)
        df[f"{labelFormasi}_Warna"] = kolomWarna
        df[f"{labelFormasi}_Posisi"] = kolomPosisi
        df[f"{labelFormasi}_KomandoNaik"] = kolomNaik
        df[f"{labelFormasi}_KomandoTurun"] = kolomTurun
    return df


def _buat_border_tipis():
    garisT = Side(style="thin", color="FFCCCCCC")
    return Border(left=garisT, right=garisT, top=garisT, bottom=garisT)


def ekspor_ke_excel(df: pd.DataFrame, daftarFormasi: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Peserta"

    fontHeaderUtama = Font(bold=True, color="FFFFFFFF", size=10, name="Calibri")
    fontHeaderSub = Font(bold=True, color="FF1a1a2e", size=9, name="Calibri")
    fillHeaderUtama = PatternFill("solid", fgColor="FF1a1a2e")
    fillHeaderSub = PatternFill("solid", fgColor="FFe0e0f0")
    alignTengah = Alignment(horizontal="center", vertical="center", wrap_text=True)
    borderTipis = _buat_border_tipis()

    jumlahFormasi = sum(1 for col in df.columns if col.endswith("_Warna"))

    ws.merge_cells("A1:A2"); ws["A1"] = "Koordinat"
    ws.merge_cells("B1:B2"); ws["B1"] = "Baris"
    ws.merge_cells("C1:C2"); ws["C1"] = "Kolom"

    KOLOM_TETAP = 3
    for idx in range(1, jumlahFormasi + 1):
        awal = KOLOM_TETAP + (idx - 1) * 4 + 1
        akhir = awal + 3
        ws.merge_cells(f"{get_column_letter(awal)}1:{get_column_letter(akhir)}1")
        sel = ws[f"{get_column_letter(awal)}1"]
        sel.value = f"Formasi {idx}"

    for idx in range(1, jumlahFormasi + 1):
        awal = KOLOM_TETAP + (idx - 1) * 4 + 1
        labelSub = ["Warna", "Posisi", "Komando Naik", "Komando Turun"]
        for j, sub in enumerate(labelSub):
            ws[f"{get_column_letter(awal + j)}2"] = sub

    for baris in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=KOLOM_TETAP + jumlahFormasi * 4):
        for sel in baris:
            if sel.row == 1:
                sel.font = fontHeaderUtama
                sel.fill = fillHeaderUtama
            else:
                sel.font = fontHeaderSub
                sel.fill = fillHeaderSub
            sel.alignment = alignTengah
            sel.border = borderTipis

    ws.freeze_panes = "D3"

    fillBergantian = PatternFill("solid", fgColor="FFF5F5FA")
    fontData = Font(size=9, name="Calibri")
    fontKoordinat = Font(size=9, name="Courier New", bold=True)

    labelKeKunci = {v["label"]: k for k, v in DEFINISI_WARNA.items()}
    fillWarna = {k: PatternFill("solid", fgColor=WARNA_EXCEL[k]) for k in DAFTAR_WARNA}
    fontWarna = {
        k: Font(size=9, name="Calibri", color=("FFFFFFFF" if k in ("hitam", "merah", "biru") else "FF000000"), bold=True)
        for k in DAFTAR_WARNA
    }
    fillBerdiri = PatternFill("solid", fgColor="FFd4edda")
    fillDuduk = PatternFill("solid", fgColor="FFf8d7da")

    daftarKolom = df.columns.tolist()
    adalahKolomWarna = [c.endswith("_Warna") for c in daftarKolom]
    adalahKolomPosisi = [c.endswith("_Posisi") for c in daftarKolom]

    for indeksBaris, dataBaris in enumerate(df.itertuples(index=False), start=3):
        barisGanjil = (indeksBaris % 2 == 0)
        for indeksKolom, nilai in enumerate(dataBaris):
            sel = ws.cell(row=indeksBaris, column=indeksKolom + 1)
            sel.value = nilai
            sel.border = borderTipis
            namaKolom = daftarKolom[indeksKolom]

            if namaKolom == "Koordinat":
                sel.font = fontKoordinat
                sel.alignment = alignTengah
            elif namaKolom in ("Baris", "Kolom"):
                sel.font = fontData
                sel.alignment = alignTengah
            elif adalahKolomWarna[indeksKolom]:
                kunciWarna = labelKeKunci.get(nilai)
                if kunciWarna:
                    sel.fill = fillWarna[kunciWarna]
                    sel.font = fontWarna[kunciWarna]
                sel.alignment = alignTengah
            elif adalahKolomPosisi[indeksKolom]:
                sel.font = fontData
                sel.alignment = alignTengah
                sel.fill = fillBerdiri if nilai == "Berdiri" else fillDuduk
            else:
                sel.font = fontData
                sel.alignment = alignTengah
                if barisGanjil and not adalahKolomWarna[indeksKolom]:
                    sel.fill = fillBergantian

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 7
    ws.column_dimensions["C"].width = 7
    for idx in range(1, jumlahFormasi + 1):
        awal = KOLOM_TETAP + (idx - 1) * 4 + 1
        ws.column_dimensions[get_column_letter(awal)].width = 10
        ws.column_dimensions[get_column_letter(awal + 1)].width = 9
        ws.column_dimensions[get_column_letter(awal + 2)].width = 13
        ws.column_dimensions[get_column_letter(awal + 3)].width = 14

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    for r in range(3, len(df) + 3):
        ws.row_dimensions[r].height = 14

    ws2 = wb.create_sheet("Ringkasan Distribusi")

    ws2["A1"] = "Ringkasan Distribusi Warna per Formasi"
    ws2["A1"].font = Font(bold=True, size=13, name="Calibri", color="FF1a1a2e")
    ws2["A1"].alignment = Alignment(horizontal="center")
    kolomAkhirHuruf = get_column_letter(jumlahFormasi + 2)
    ws2.merge_cells(f"A1:{kolomAkhirHuruf}1")

    barisSub = 2
    ws2.cell(barisSub, 1, "Warna")
    for i in range(1, jumlahFormasi + 1):
        ws2.cell(barisSub, i + 1, f"Formasi {i}")
    ws2.cell(barisSub, jumlahFormasi + 2, "Total")

    for c in range(1, jumlahFormasi + 3):
        sel = ws2.cell(barisSub, c)
        sel.font = fontHeaderUtama
        sel.fill = fillHeaderUtama
        sel.alignment = alignTengah
        sel.border = borderTipis

    fillBergantian2 = PatternFill("solid", fgColor="FFF5F5FA")
    for wi, kunciW in enumerate(DAFTAR_WARNA):
        b = barisSub + 1 + wi
        ws2.cell(b, 1, DEFINISI_WARNA[kunciW]["label"])
        ws2.cell(b, 1).fill = PatternFill("solid", fgColor=WARNA_EXCEL[kunciW])
        warnaFont = ("FFFFFFFF" if kunciW in ("hitam", "merah", "biru") else "FF000000")
        ws2.cell(b, 1).font = Font(bold=True, size=9, color=warnaFont, name="Calibri")
        ws2.cell(b, 1).alignment = alignTengah
        ws2.cell(b, 1).border = borderTipis

        totalBaris = 0
        for fi, dataF in enumerate(daftarFormasi):
            jml = dataF["distribusi"].get(kunciW, 0)
            totalBaris += jml
            selData = ws2.cell(b, fi + 2, jml)
            selData.font = fontData
            selData.alignment = alignTengah
            selData.border = borderTipis
            if wi % 2 == 1:
                selData.fill = fillBergantian2

        selTotal = ws2.cell(b, jumlahFormasi + 2, totalBaris)
        selTotal.font = Font(bold=True, size=9, name="Calibri")
        selTotal.alignment = alignTengah
        selTotal.border = borderTipis

    barisTotalAkhir = barisSub + len(DAFTAR_WARNA) + 1
    ws2.cell(barisTotalAkhir, 1, "TOTAL")
    ws2.cell(barisTotalAkhir, 1).font = Font(bold=True, size=9)
    ws2.cell(barisTotalAkhir, 1).alignment = alignTengah
    ws2.cell(barisTotalAkhir, 1).border = borderTipis
    grandTotal = 0
    for fi, dataF in enumerate(daftarFormasi):
        s = sum(dataF["distribusi"].values())
        grandTotal += s
        c = ws2.cell(barisTotalAkhir, fi + 2, s)
        c.font = Font(bold=True, size=9)
        c.alignment = alignTengah
        c.border = borderTipis
    ws2.cell(barisTotalAkhir, jumlahFormasi + 2, grandTotal).font = Font(bold=True, size=9)
    ws2.cell(barisTotalAkhir, jumlahFormasi + 2).alignment = alignTengah
    ws2.cell(barisTotalAkhir, jumlahFormasi + 2).border = borderTipis

    ws2.column_dimensions["A"].width = 12
    for i in range(2, jumlahFormasi + 3):
        ws2.column_dimensions[get_column_letter(i)].width = 11

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
