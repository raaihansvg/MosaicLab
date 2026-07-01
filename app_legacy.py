import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import numpy as np
import pandas as pd
from PIL import Image, ImageTk
import io
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import threading
import warnings

warnings.filterwarnings("ignore")

GRID_SIZE    = 60
N_COLORS     = 6
N_FORMATIONS = 10

WARNA_DEFINISI = {
    "hitam":  {"rgb": (0,   0,   0),   "hex": "#000000", "label": "Hitam"},
    "putih":  {"rgb": (255, 255, 255), "hex": "#FFFFFF", "label": "Putih"},
    "oren":   {"rgb": (255, 140,  0),  "hex": "#FF8C00", "label": "Oren"},
    "merah":  {"rgb": (220,  20, 60),  "hex": "#DC143C", "label": "Merah"},
    "kuning": {"rgb": (255, 215,  0),  "hex": "#FFD700", "label": "Kuning"},
    "biru":   {"rgb": (30,  144, 255), "hex": "#1E90FF", "label": "Biru"},
}

WARNA_KEYS = list(WARNA_DEFINISI.keys())
WARNA_RGB  = np.array([v["rgb"] for v in WARNA_DEFINISI.values()], dtype=np.float32)

BACKGROUND_COLORS = {"putih", "hitam"}

EXCEL_FILLS = {
    "hitam":  "FF000000",
    "putih":  "FFFFFFFF",
    "oren":   "FFFF8C00",
    "merah":  "FFDC143C",
    "kuning": "FFFFD700",
    "biru":   "FF1E90FF",
}

PRIMARY      = "#E8621A"
PRIMARY_DARK = "#C4511A"
SECONDARY    = "#F5A623"
BG_BASE      = "#FAFAF8"
BG_SURFACE   = "#FFFFFF"
BG_SIDEBAR   = "#1C1917"
TEXT_PRIMARY = "#1C1917"
TEXT_MUTED   = "#78716C"
TEXT_INVERSE = "#FAFAF8"
BORDER       = "#E7E5E4"
CLR_SUCCESS  = "#16A34A"
CLR_WARNING  = "#D97706"
CLR_ERROR    = "#DC2626"

def load_image(uploaded_file) -> np.ndarray:
    try:
        img = Image.open(uploaded_file).convert("RGB")
        return np.array(img)
    except Exception as e:
        raise ValueError(f"Gagal memuat gambar: {e}")


def resize_to_grid(img_array: np.ndarray, size: int = GRID_SIZE) -> np.ndarray:
    pil_img = Image.fromarray(img_array)
    resized  = pil_img.resize((size, size), Image.LANCZOS)
    return np.array(resized)


def _pixels_to_hsv(rgb_array: np.ndarray) -> np.ndarray:
    rgb  = rgb_array.astype(np.float32) / 255.0
    r, g, b = rgb[:, 0], rgb[:, 1], rgb[:, 2]
    cmax  = np.maximum(np.maximum(r, g), b)
    cmin  = np.minimum(np.minimum(r, g), b)
    delta = cmax - cmin

    h = np.zeros(len(r), dtype=np.float32)
    mask = delta > 0
    m = mask & (cmax == r)
    h[m] = 60.0 * (((g[m] - b[m]) / delta[m]) % 6)
    m = mask & (cmax == g)
    h[m] = 60.0 * (((b[m] - r[m]) / delta[m]) + 2)
    m = mask & (cmax == b)
    h[m] = 60.0 * (((r[m] - g[m]) / delta[m]) + 4)
    h = h % 360.0

    s = np.where(cmax > 0, delta / cmax, 0.0)
    v = cmax
    return np.stack([h, s, v], axis=1)


def map_pixel_to_color_hsv(rgb_array: np.ndarray) -> np.ndarray:
    hsv     = _pixels_to_hsv(rgb_array)
    h, s, v = hsv[:, 0], hsv[:, 1], hsv[:, 2]
    n       = len(h)

    IDX    = {k: i for i, k in enumerate(WARNA_KEYS)}
    result = np.full(n, -1, dtype=np.int32)


    result[v < 0.40] = IDX["hitam"]


    mask = (result == -1) & (v > 0.72) & (s < 0.40)
    result[mask] = IDX["putih"]


    mask = (result == -1) & (s < 0.22)
    result[mask & (v <= 0.55)] = IDX["hitam"]
    result[mask & (v >  0.55)] = IDX["putih"]


    mask = (result == -1) & (h >= 8) & (h < 35) & (s >= 0.55) & (v >= 0.50)
    result[mask] = IDX["oren"]


    mask = (result == -1) & (h >= 35) & (h < 80) & (s >= 0.25) & (v >= 0.42)
    result[mask] = IDX["kuning"]


    mask = (result == -1) & (h >= 180) & (h < 270) & (s >= 0.50) & (v >= 0.50)
    result[mask] = IDX["biru"]


    mask = (result == -1) & ((h < 8) | (h >= 315)) & (s >= 0.55) & (v >= 0.50)
    result[mask] = IDX["merah"]


    mask = (result == -1) & (h >= 270) & (h < 315) & (s >= 0.55) & (v >= 0.50)
    result[mask] = IDX["merah"]


    unresolved = result == -1
    if unresolved.any():
        px_u  = rgb_array[unresolved].astype(np.float32)
        diff  = px_u[:, np.newaxis, :] - WARNA_RGB[np.newaxis, :, :]
        dist  = np.sum(diff ** 2, axis=2)
        best  = np.argmin(dist, axis=1)
        v_u   = v[unresolved]
        s_u   = s[unresolved]

        CHROMATIC  = {IDX["oren"], IDX["merah"], IDX["kuning"], IDX["biru"]}
        final_best = best.copy()
        for pi in range(len(px_u)):
            if best[pi] in CHROMATIC:
                d_chroma = dist[pi, best[pi]]
                d_hitam  = dist[pi, IDX["hitam"]]
                d_putih  = dist[pi, IDX["putih"]]


                if v_u[pi] < 0.55 and d_hitam < 2.0 * d_chroma:
                    final_best[pi] = IDX["hitam"]


                elif v_u[pi] >= 0.50 and s_u[pi] < 0.50 and d_putih < 2.5 * d_chroma:
                    final_best[pi] = IDX["putih"]

        result[unresolved] = final_best

    return result


def preprocess_image_for_grid(img_array: np.ndarray) -> np.ndarray:
    from PIL import ImageFilter
    INTER_SIZE = 500

    pil_img = Image.fromarray(img_array)


    mid = pil_img.resize((INTER_SIZE, INTER_SIZE), Image.LANCZOS)


    mid = mid.filter(ImageFilter.UnsharpMask(radius=2, percent=250, threshold=2))


    final = mid.resize((GRID_SIZE, GRID_SIZE), Image.BOX)

    return np.array(final)


def map_to_predefined_colors(img_array: np.ndarray) -> np.ndarray:
    pixels = img_array.reshape(-1, 3)
    color_indices = map_pixel_to_color_hsv(pixels)
    return color_indices.reshape(GRID_SIZE, GRID_SIZE)


def process_formation_image(uploaded_file) -> dict:
    raw        = load_image(uploaded_file)


    preprocessed = preprocess_image_for_grid(raw)
    color_grid   = map_to_predefined_colors(preprocessed)

    color_names = np.array(WARNA_KEYS, dtype=object)[color_grid]
    preview_rgb = WARNA_RGB[color_grid].astype(np.uint8)

    distribution = {}
    for i, key in enumerate(WARNA_KEYS):
        distribution[key] = int(np.sum(color_grid == i))

    return {
        "color_grid":   color_grid,
        "color_names":  color_names,
        "preview_rgb":  preview_rgb,
        "distribution": distribution,
    }


def render_grid_preview(
    formation_data: dict,
    formation_label: str,
    show_grid_lines: bool = True,
) -> plt.Figure:
    preview = formation_data["preview_rgb"]

    fig, ax = plt.subplots(1, 1, figsize=(10, 10), facecolor=BG_BASE)
    ax.set_facecolor(BG_SURFACE)
    ax.imshow(preview, interpolation="nearest", aspect="equal")
    ax.set_title(
        f"Grid Formasi — {formation_label}",
        color=TEXT_PRIMARY, fontsize=13, fontweight="bold", pad=12,
        fontfamily="Segoe UI",
    )
    ax.set_xlabel("Kolom (1-60)", color=TEXT_MUTED, fontsize=11)
    ax.set_ylabel("Baris (1-60)", color=TEXT_MUTED, fontsize=11)
    ax.tick_params(colors=TEXT_MUTED, labelsize=9)
    for spine in ax.spines.values():
        spine.set_edgecolor(BORDER)

    if show_grid_lines:
        for i in range(0, GRID_SIZE + 1, 1):
            ax.axhline(i - 0.5, color=BORDER, linewidth=0.3)
            ax.axvline(i - 0.5, color=BORDER, linewidth=0.3)
        for i in range(0, GRID_SIZE + 1, 10):
            ax.axhline(i - 0.5, color=TEXT_MUTED, linewidth=0.7, alpha=0.5)
            ax.axvline(i - 0.5, color=TEXT_MUTED, linewidth=0.7, alpha=0.5)

    ticks = list(range(0, GRID_SIZE, 5))
    ax.set_xticks(ticks)
    ax.set_xticklabels([t + 1 for t in ticks], fontsize=8)
    ax.set_yticks(ticks)
    ax.set_yticklabels([t + 1 for t in ticks], fontsize=8)

    plt.tight_layout(pad=2)
    return fig


def generate_participant_base() -> pd.DataFrame:
    rows, cols = [], []
    for r in range(1, GRID_SIZE + 1):
        for c in range(1, GRID_SIZE + 1):
            rows.append(r)
            cols.append(c)
    koordinat = [f"({r},{c})" for r, c in zip(rows, cols)]
    return pd.DataFrame({"Koordinat": koordinat, "Baris": rows, "Kolom": cols})


def determine_posisi(color_name: str, bg_color: str = None) -> tuple:
    effective_bg = BACKGROUND_COLORS.copy()
    if bg_color:
        effective_bg.add(bg_color)
    if color_name in effective_bg:
        return ("Duduk", "-", "tidak turun")
    else:
        return ("Berdiri", "Mulai", "Turun")


def build_participant_dataframe(formations_data: list, bg_color: str = None) -> pd.DataFrame:
    df = generate_participant_base()
    for i, fdata in enumerate(formations_data, start=1):
        label  = f"F{i}"
        cn_arr = fdata["color_names"].reshape(-1)
        warna_col, posisi_col, naik_col, turun_col = [], [], [], []
        for cn in cn_arr:
            posisi, naik, turun = determine_posisi(cn, bg_color=bg_color)
            warna_col.append(WARNA_DEFINISI[cn]["label"])
            posisi_col.append(posisi)
            naik_col.append(naik)
            turun_col.append(turun)
        df[f"{label}_Warna"]        = warna_col
        df[f"{label}_Posisi"]       = posisi_col
        df[f"{label}_KomandoNaik"]  = naik_col
        df[f"{label}_KomandoTurun"] = turun_col
    return df


def _make_thin_border():
    thin = Side(style="thin", color="FFCCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def export_to_excel(df: pd.DataFrame, formations_data: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Peserta"

    hdr_font_main = Font(bold=True, color="FFFFFFFF", size=10, name="Calibri")
    hdr_font_sub  = Font(bold=True, color="FF1a1a2e", size=9,  name="Calibri")
    hdr_fill_main = PatternFill("solid", fgColor="FF1a1a2e")
    hdr_fill_sub  = PatternFill("solid", fgColor="FFe0e0f0")
    center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border   = _make_thin_border()

    n_formations = sum(1 for col in df.columns if col.endswith("_Warna"))

    ws.merge_cells("A1:A2"); ws["A1"] = "Koordinat"
    ws.merge_cells("B1:B2"); ws["B1"] = "Baris"
    ws.merge_cells("C1:C2"); ws["C1"] = "Kolom"

    fixed_cols = 3
    for idx in range(1, n_formations + 1):
        start = fixed_cols + (idx - 1) * 4 + 1
        end   = start + 3
        ws.merge_cells(f"{get_column_letter(start)}1:{get_column_letter(end)}1")
        cell = ws[f"{get_column_letter(start)}1"]
        cell.value = f"Formasi {idx}"

    for idx in range(1, n_formations + 1):
        start = fixed_cols + (idx - 1) * 4 + 1
        sub_labels = ["Warna", "Posisi", "Komando Naik", "Komando Turun"]
        for j, sub in enumerate(sub_labels):
            ws[f"{get_column_letter(start + j)}2"] = sub

    for row in ws.iter_rows(min_row=1, max_row=2,
                             min_col=1, max_col=fixed_cols + n_formations * 4):
        for cell in row:
            if cell.row == 1:
                cell.font = hdr_font_main
                cell.fill = hdr_fill_main
            else:
                cell.font = hdr_font_sub
                cell.fill = hdr_fill_sub
            cell.alignment = center_align
            cell.border    = thin_border

    ws.freeze_panes = "D3"

    col_map       = {col: i + 1 for i, col in enumerate(df.columns)}
    row_fill_alt  = PatternFill("solid", fgColor="FFF5F5FA")
    data_font     = Font(size=9, name="Calibri")
    coord_font    = Font(size=9, name="Courier New", bold=True)

    for row_idx, (_, row) in enumerate(df.iterrows(), start=3):
        is_alt = (row_idx % 2 == 0)
        for col_name, col_excel in col_map.items():
            cell = ws.cell(row=row_idx, column=col_excel)
            cell.value  = row[col_name]
            cell.border = thin_border
            if col_name == "Koordinat":
                cell.font      = coord_font
                cell.alignment = center_align
            elif col_name in ("Baris", "Kolom"):
                cell.font      = data_font
                cell.alignment = center_align
            elif col_name.endswith("_Warna"):
                warna_key = next(
                    (k for k, v in WARNA_DEFINISI.items() if v["label"] == row[col_name]),
                    None,
                )
                if warna_key:
                    cell.fill = PatternFill("solid", fgColor=EXCEL_FILLS[warna_key])
                    txt_color = ("FFFFFFFF"
                                 if warna_key in ("hitam", "merah", "biru")
                                 else "FF000000")
                    cell.font = Font(size=9, name="Calibri", color=txt_color, bold=True)
                cell.alignment = center_align
            elif col_name.endswith("_Posisi"):
                cell.font      = data_font
                cell.alignment = center_align
                cell.fill = (PatternFill("solid", fgColor="FFd4edda")
                             if row[col_name] == "Berdiri"
                             else PatternFill("solid", fgColor="FFf8d7da"))
            else:
                cell.font      = data_font
                cell.alignment = center_align
                if is_alt and not col_name.endswith("_Warna"):
                    cell.fill = row_fill_alt

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 7
    ws.column_dimensions["C"].width = 7
    for idx in range(1, n_formations + 1):
        start = fixed_cols + (idx - 1) * 4 + 1
        ws.column_dimensions[get_column_letter(start)].width     = 10
        ws.column_dimensions[get_column_letter(start + 1)].width = 9
        ws.column_dimensions[get_column_letter(start + 2)].width = 13
        ws.column_dimensions[get_column_letter(start + 3)].width = 14

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    for r in range(3, len(df) + 3):
        ws.row_dimensions[r].height = 14


    ws2 = wb.create_sheet("Ringkasan Distribusi")

    ws2["A1"] = "Ringkasan Distribusi Warna per Formasi"
    ws2["A1"].font      = Font(bold=True, size=13, name="Calibri", color="FF1a1a2e")
    ws2["A1"].alignment = Alignment(horizontal="center")
    last_col_letter = get_column_letter(n_formations + 2)
    ws2.merge_cells(f"A1:{last_col_letter}1")

    sub_row = 2
    ws2.cell(sub_row, 1, "Warna")
    for i in range(1, n_formations + 1):
        ws2.cell(sub_row, i + 1, f"Formasi {i}")
    ws2.cell(sub_row, n_formations + 2, "Total")

    for c in range(1, n_formations + 3):
        cell = ws2.cell(sub_row, c)
        cell.font      = hdr_font_main
        cell.fill      = hdr_fill_main
        cell.alignment = center_align
        cell.border    = thin_border

    row_fill_alt2 = PatternFill("solid", fgColor="FFF5F5FA")
    for wi, wkey in enumerate(WARNA_KEYS):
        r = sub_row + 1 + wi
        ws2.cell(r, 1, WARNA_DEFINISI[wkey]["label"])
        ws2.cell(r, 1).fill = PatternFill("solid", fgColor=EXCEL_FILLS[wkey])
        txt_c = ("FFFFFFFF" if wkey in ("hitam", "merah", "biru") else "FF000000")
        ws2.cell(r, 1).font      = Font(bold=True, size=9, color=txt_c, name="Calibri")
        ws2.cell(r, 1).alignment = center_align
        ws2.cell(r, 1).border    = thin_border

        row_total = 0
        for fi, fdata in enumerate(formations_data):
            cnt = fdata["distribution"].get(wkey, 0)
            row_total += cnt
            cell = ws2.cell(r, fi + 2, cnt)
            cell.font      = data_font
            cell.alignment = center_align
            cell.border    = thin_border
            if wi % 2 == 1:
                cell.fill = row_fill_alt2

        cell_total = ws2.cell(r, n_formations + 2, row_total)
        cell_total.font      = Font(bold=True, size=9, name="Calibri")
        cell_total.alignment = center_align
        cell_total.border    = thin_border

    tot_row = sub_row + len(WARNA_KEYS) + 1
    ws2.cell(tot_row, 1, "TOTAL")
    ws2.cell(tot_row, 1).font      = Font(bold=True, size=9)
    ws2.cell(tot_row, 1).alignment = center_align
    ws2.cell(tot_row, 1).border    = thin_border
    grand = 0
    for fi, fdata in enumerate(formations_data):
        s = sum(fdata["distribution"].values())
        grand += s
        c = ws2.cell(tot_row, fi + 2, s)
        c.font      = Font(bold=True, size=9)
        c.alignment = center_align
        c.border    = thin_border
    ws2.cell(tot_row, n_formations + 2, grand).font = Font(bold=True, size=9)
    ws2.cell(tot_row, n_formations + 2).alignment = center_align
    ws2.cell(tot_row, n_formations + 2).border = thin_border

    ws2.column_dimensions["A"].width = 12
    for i in range(2, n_formations + 3):
        ws2.column_dimensions[get_column_letter(i)].width = 11

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


ALLOWED_EXT = {".jpg", ".jpeg", ".png", ".webp"}

class ODMApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("ODM Undip 2026 — Mozaik Formation Software")
        self.minsize(1280, 800)
        self.configure(bg=BG_BASE)

        self.uploaded_paths = {}
        self.formations_data = {}
        self.excel_bytes = None
        self.show_grid_var = tk.BooleanVar(value=True)
        self.bg_color_var = tk.StringVar(value="putih")
        self._thumb_refs = {}
        self._canvas_refs = {}
        self._active_tab = None
        self._df_preview = None

        self._build_styles()
        self._build_layout()
        self._refresh_sidebar_state()

    def _build_styles(self):
        s = ttk.Style(self)
        s.theme_use("clam")
        s.configure(".", background=BG_BASE, foreground=TEXT_PRIMARY, font=("Segoe UI", 11))
        s.configure("TFrame", background=BG_BASE)
        s.configure("Sidebar.TFrame", background=BG_SIDEBAR)
        s.configure("Surface.TFrame", background=BG_SURFACE)
        s.configure("TLabel", background=BG_BASE, foreground=TEXT_PRIMARY, font=("Segoe UI", 11))
        s.configure("Treeview", background=BG_SURFACE, foreground=TEXT_PRIMARY,
                     fieldbackground=BG_SURFACE, rowheight=24, font=("Segoe UI", 9))
        s.configure("Treeview.Heading", background=BG_BASE, foreground=TEXT_PRIMARY,
                     font=("Segoe UI", 9, "bold"), relief="flat")
        s.map("Treeview", background=[("selected", "#FFF7ED")], foreground=[("selected", TEXT_PRIMARY)])
        s.configure("TScrollbar", background=BG_BASE, troughcolor=BORDER, arrowcolor=TEXT_MUTED)
        s.configure("TCombobox", fieldbackground="#292524", background="#292524",
                     foreground=TEXT_INVERSE, arrowcolor=TEXT_INVERSE,
                     selectbackground=PRIMARY, selectforeground=TEXT_INVERSE)
        self.option_add("*TCombobox*Listbox.background", "#292524")
        self.option_add("*TCombobox*Listbox.foreground", TEXT_INVERSE)
        self.option_add("*TCombobox*Listbox.selectBackground", PRIMARY)
        self.option_add("*TCombobox*Listbox.selectForeground", TEXT_INVERSE)
        self.style = s

    def _build_layout(self):
        self.columnconfigure(0, weight=0, minsize=280)
        self.columnconfigure(1, weight=1)
        self.rowconfigure(0, weight=1)
        self._build_sidebar()
        self._build_main()


    def _build_sidebar(self):
        sb = tk.Frame(self, bg=BG_SIDEBAR, width=280)
        sb.grid(row=0, column=0, sticky="nsew")
        sb.grid_propagate(False)
        sb.columnconfigure(0, weight=1)
        sb.rowconfigure(1, weight=1)


        hdr = tk.Frame(sb, bg=BG_SIDEBAR)
        hdr.pack(fill="x", padx=20, pady=(20, 0))
        tk.Label(hdr, text="ODM Undip 2026", bg=BG_SIDEBAR, fg=TEXT_INVERSE,
                 font=("Segoe UI", 16, "bold")).pack(anchor="w")
        tk.Label(hdr, text="Mozaik Formation Software", bg=BG_SIDEBAR, fg=TEXT_MUTED,
                 font=("Segoe UI", 10)).pack(anchor="w")


        tk.Frame(sb, bg="#292524", height=1).pack(fill="x", padx=16, pady=12)


        tk.Label(sb, text="Gambar Formasi", bg=BG_SIDEBAR, fg=TEXT_INVERSE,
                 font=("Segoe UI", 11, "bold")).pack(padx=20, anchor="w")


        slot_canvas = tk.Canvas(sb, bg=BG_SIDEBAR, bd=0, highlightthickness=0)
        slot_scroll = ttk.Scrollbar(sb, orient="vertical", command=slot_canvas.yview)
        slot_canvas.configure(yscrollcommand=slot_scroll.set)
        slot_scroll.pack(side="right", fill="y")
        slot_canvas.pack(fill="both", expand=True, padx=(0, 0))

        inner = tk.Frame(slot_canvas, bg=BG_SIDEBAR)
        win_id = slot_canvas.create_window((0, 0), window=inner, anchor="nw")
        inner.bind("<Configure>", lambda e: slot_canvas.configure(scrollregion=slot_canvas.bbox("all")))
        slot_canvas.bind("<Configure>", lambda e: slot_canvas.itemconfig(win_id, width=e.width))

        def _mw(event):
            slot_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        slot_canvas.bind_all("<MouseWheel>", _mw)

        self._slot_fname_labels = {}
        self._slot_dots = {}
        self._slot_x_btns = {}
        self._slot_err_labels = {}

        for i in range(1, N_FORMATIONS + 1):
            container = tk.Frame(inner, bg=BG_SIDEBAR)
            container.pack(fill="x", padx=16, pady=3)

            row = tk.Frame(container, bg=BG_SIDEBAR)
            row.pack(fill="x")


            tk.Label(row, text=f"F{i:02d}", bg=BG_SIDEBAR, fg=TEXT_MUTED,
                     font=("Segoe UI", 9), width=4, anchor="w").pack(side="left")


            fl = tk.Label(row, text="Belum diunggah", bg=BG_SIDEBAR, fg="#A8A29E",
                          font=("Segoe UI", 9), anchor="w", cursor="hand2")
            fl.pack(side="left", fill="x", expand=True)
            fl.bind("<Button-1>", lambda e, idx=i: self._pick_image(idx))
            self._slot_fname_labels[i] = fl


            dot = tk.Canvas(row, width=12, height=12, bg=BG_SIDEBAR, highlightthickness=0)
            dot.create_oval(2, 2, 10, 10, fill=BORDER, outline="", tags="dot")
            dot.pack(side="left", padx=4)
            self._slot_dots[i] = dot


            xb = tk.Label(row, text="X", bg=BG_SIDEBAR, fg="#57534E",
                          font=("Segoe UI", 8), cursor="hand2", padx=4)
            xb.pack(side="left")
            xb.bind("<Button-1>", lambda e, idx=i: self._remove_formation(idx))
            xb.bind("<Enter>", lambda e, w=xb: w.config(fg=CLR_ERROR))
            xb.bind("<Leave>", lambda e, w=xb: w.config(fg="#57534E"))
            self._slot_x_btns[i] = xb


            err = tk.Label(container, text="", bg=BG_SIDEBAR, fg=CLR_ERROR,
                           font=("Segoe UI", 8), anchor="w")
            self._slot_err_labels[i] = err


        self._upload_count_lbl = tk.Label(sb, text="0 / 10 diunggah",
                                          bg=BG_SIDEBAR, fg=TEXT_MUTED,
                                          font=("Segoe UI", 9))
        self._upload_count_lbl.pack(padx=20, pady=(6, 0), anchor="w")


        tk.Frame(sb, bg="#292524", height=1).pack(fill="x", padx=16, pady=10)


        tk.Label(sb, text="Pengaturan", bg=BG_SIDEBAR, fg=TEXT_INVERSE,
                 font=("Segoe UI", 11, "bold")).pack(padx=20, anchor="w")


        tog_row = tk.Frame(sb, bg=BG_SIDEBAR)
        tog_row.pack(fill="x", padx=20, pady=6)
        tk.Label(tog_row, text="Garis grid", bg=BG_SIDEBAR, fg="#D6D3D1",
                 font=("Segoe UI", 10)).pack(side="left")
        self._toggle_canvas = self._create_toggle(tog_row, self.show_grid_var)
        self._toggle_canvas.pack(side="right")


        tk.Label(sb, text="Warna background", bg=BG_SIDEBAR, fg="#D6D3D1",
                 font=("Segoe UI", 10)).pack(padx=20, anchor="w", pady=(4, 2))
        bg_opts = {v["label"]: k for k, v in WARNA_DEFINISI.items() if k in BACKGROUND_COLORS}
        self._bg_display_map = bg_opts
        self._bg_combo_var = tk.StringVar(value="Putih")
        combo = ttk.Combobox(sb, textvariable=self._bg_combo_var,
                             values=list(bg_opts.keys()), state="readonly", width=16)
        combo.pack(padx=20, anchor="w")


        self._process_tip = tk.Label(sb, text="Upload minimal 1 gambar untuk memulai.",
                                     bg=BG_SIDEBAR, fg=TEXT_MUTED, font=("Segoe UI", 8),
                                     wraplength=240)
        self._process_tip.pack(padx=20, pady=(12, 2), anchor="w")


        self._process_btn = tk.Label(sb, text="Proses Semua Gambar",
                                     bg=PRIMARY, fg=TEXT_INVERSE,
                                     font=("Segoe UI", 12, "bold"),
                                     cursor="hand2", padx=16, pady=10)
        self._process_btn.pack(fill="x", padx=16, pady=(4, 20))
        self._process_btn.bind("<Button-1>", lambda e: self._start_processing())
        self._process_btn.bind("<Enter>", lambda e: self._process_btn.config(bg=PRIMARY_DARK))
        self._process_btn.bind("<Leave>", lambda e: self._process_btn.config(
            bg=PRIMARY if self.uploaded_paths else "#A8A29E"))

    def _create_toggle(self, parent, variable):
        W, H = 44, 24
        c = tk.Canvas(parent, width=W, height=H, bg=BG_SIDEBAR, highlightthickness=0, cursor="hand2")
        def draw():
            c.delete("all")
            on = variable.get()
            bg = PRIMARY if on else "#57534E"
            r = H // 2
            c.create_oval(0, 0, H, H, fill=bg, outline="")
            c.create_oval(W - H, 0, W, H, fill=bg, outline="")
            c.create_rectangle(r, 0, W - r, H, fill=bg, outline="")
            kx = W - r if on else r
            c.create_oval(kx - 8, 4, kx + 8, H - 4, fill="white", outline="")
        def toggle(e=None):
            variable.set(not variable.get())
            draw()
        c.bind("<Button-1>", toggle)
        draw()
        return c


    def _build_main(self):
        main = tk.Frame(self, bg=BG_BASE)
        main.grid(row=0, column=1, sticky="nsew")
        main.columnconfigure(0, weight=1)
        main.rowconfigure(1, weight=1)


        topbar = tk.Frame(main, bg=BG_SURFACE, height=44)
        topbar.grid(row=0, column=0, sticky="ew")
        topbar.grid_propagate(False)
        topbar.columnconfigure(0, weight=1)
        tk.Label(topbar, text="Dashboard / Preview Formasi", bg=BG_SURFACE,
                 fg=TEXT_MUTED, font=("Segoe UI", 10)).pack(side="left", padx=20)
        tk.Label(topbar, text="Grid 60x60  |  3.600 Peserta  |  6 Warna",
                 bg=BG_SURFACE, fg=TEXT_MUTED, font=("Segoe UI", 10)).pack(side="right", padx=20)
        tk.Frame(topbar, bg=BORDER, height=1).place(relx=0, rely=1.0, relwidth=1.0, anchor="sw")


        outer = tk.Frame(main, bg=BG_BASE)
        outer.grid(row=1, column=0, sticky="nsew")
        outer.columnconfigure(0, weight=1)
        outer.rowconfigure(0, weight=1)

        self._main_canvas = tk.Canvas(outer, bg=BG_BASE, bd=0, highlightthickness=0)
        vscroll = ttk.Scrollbar(outer, orient="vertical", command=self._main_canvas.yview)
        self._main_canvas.configure(yscrollcommand=vscroll.set)
        vscroll.grid(row=0, column=1, sticky="ns")
        self._main_canvas.grid(row=0, column=0, sticky="nsew")

        self._content = tk.Frame(self._main_canvas, bg=BG_BASE)
        self._content_id = self._main_canvas.create_window((0, 0), window=self._content, anchor="nw")
        self._content.bind("<Configure>",
            lambda e: self._main_canvas.configure(scrollregion=self._main_canvas.bbox("all")))
        self._main_canvas.bind("<Configure>",
            lambda e: self._main_canvas.itemconfig(self._content_id, width=e.width))

        def _scroll_main(event):
            self._main_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        self._main_canvas.bind("<MouseWheel>", _scroll_main)
        self._content.bind("<MouseWheel>", _scroll_main)

        self._build_content_area()

    def _build_content_area(self):
        c = self._content


        self._metrics_frame = tk.Frame(c, bg=BG_BASE)
        self._metrics_frame.pack(fill="x", padx=24, pady=(16, 8))
        self._build_metrics()


        self._step_frame = tk.Frame(c, bg=BG_BASE)
        self._step_frame.pack(fill="x", padx=24, pady=8)
        self._build_step_indicator()


        self._proc_frame = tk.Frame(c, bg=BG_BASE)
        self._proc_label = tk.Label(self._proc_frame, text="", bg=BG_BASE, fg=TEXT_MUTED,
                                    font=("Segoe UI", 10))
        self._proc_label.pack(anchor="w")
        self._proc_bar_canvas = tk.Canvas(self._proc_frame, height=6, bg=BORDER,
                                          highlightthickness=0)
        self._proc_bar_canvas.pack(fill="x", pady=(4, 0))
        self._proc_bar_fill = None


        tk.Label(c, text="Preview Formasi", bg=BG_BASE, fg=TEXT_PRIMARY,
                 font=("Segoe UI", 13, "bold")).pack(padx=24, pady=(12, 4), anchor="w")

        self._tab_row = tk.Frame(c, bg=BG_BASE)
        self._tab_row.pack(fill="x", padx=24, pady=(0, 2))
        self._tab_labels = {}
        self._tab_indicators = {}
        self._build_tab_row()


        self._formation_frame = tk.Frame(c, bg=BG_BASE)
        self._formation_frame.pack(fill="both", expand=True, padx=24, pady=(0, 8))
        self._show_empty_state()


        tk.Label(c, text="Ringkasan Distribusi Warna", bg=BG_BASE, fg=TEXT_PRIMARY,
                 font=("Segoe UI", 13, "bold")).pack(padx=24, pady=(8, 4), anchor="w")
        self._dist_frame = tk.Frame(c, bg=BG_BASE)
        self._dist_frame.pack(fill="x", padx=24, pady=(0, 8))
        self._dist_tree = None


        exp_card = tk.Frame(c, bg=BG_SURFACE, highlightbackground=BORDER, highlightthickness=1)
        exp_card.pack(fill="x", padx=24, pady=8)

        tk.Label(exp_card, text="Export Data Excel", bg=BG_SURFACE, fg=TEXT_PRIMARY,
                 font=("Segoe UI", 13, "bold")).pack(padx=16, pady=(12, 4), anchor="w")
        tk.Label(exp_card, text="Generate file Excel berisi 3.600 baris data peserta dengan "
                 "koordinat, warna, posisi, dan komando untuk setiap formasi.",
                 bg=BG_SURFACE, fg=TEXT_MUTED, font=("Segoe UI", 10),
                 wraplength=700, justify="left").pack(padx=16, anchor="w")

        btn_row = tk.Frame(exp_card, bg=BG_SURFACE)
        btn_row.pack(padx=16, pady=8, anchor="w")


        self._gen_btn = tk.Label(btn_row, text="Generate Excel", bg=BG_SURFACE,
                                 fg=PRIMARY, font=("Segoe UI", 10, "bold"),
                                 cursor="hand2", padx=16, pady=6,
                                 highlightbackground=PRIMARY, highlightthickness=2)
        self._gen_btn.pack(side="left", padx=(0, 8))
        self._gen_btn.bind("<Button-1>", lambda e: self._generate_excel())
        self._gen_btn.bind("<Enter>", lambda e: self._gen_btn.config(bg="#FFF7ED"))
        self._gen_btn.bind("<Leave>", lambda e: self._gen_btn.config(bg=BG_SURFACE))


        self._dl_btn = tk.Label(btn_row, text="Unduh File Excel", bg="#D6D3D1",
                                fg=BG_SURFACE, font=("Segoe UI", 10, "bold"),
                                padx=16, pady=6)
        self._dl_btn.pack(side="left")
        self._dl_btn_enabled = False

        self._excel_status = tk.Label(exp_card, text="", bg=BG_SURFACE, fg=TEXT_MUTED,
                                      font=("Segoe UI", 9))
        self._excel_status.pack(padx=16, pady=(0, 12), anchor="w")


        tk.Label(c, text="Pratinjau 10 baris pertama", bg=BG_BASE, fg=TEXT_PRIMARY,
                 font=("Segoe UI", 13, "bold")).pack(padx=24, pady=(8, 4), anchor="w")
        self._preview_frame = tk.Frame(c, bg=BG_BASE)
        self._preview_frame.pack(fill="x", padx=24, pady=(0, 24))
        self._preview_tree = None

    def _build_metrics(self):
        for w in self._metrics_frame.winfo_children():
            w.destroy()
        metrics = [
            ("Formasi Diproses", f"{len(self.formations_data)} / {N_FORMATIONS}"),
            ("Total Peserta", f"{GRID_SIZE * GRID_SIZE:,}"),
            ("Ukuran Grid", f"{GRID_SIZE} x {GRID_SIZE}"),
        ]
        for i, (title, value) in enumerate(metrics):
            card = tk.Frame(self._metrics_frame, bg=BG_SURFACE,
                            highlightbackground=BORDER, highlightthickness=1)
            card.grid(row=0, column=i, padx=6, pady=4, sticky="ew")
            self._metrics_frame.columnconfigure(i, weight=1)
            accent = tk.Frame(card, bg=PRIMARY, width=4)
            accent.pack(side="left", fill="y")
            inner = tk.Frame(card, bg=BG_SURFACE)
            inner.pack(side="left", fill="both", expand=True, padx=14, pady=10)
            tk.Label(inner, text=value, bg=BG_SURFACE, fg=TEXT_PRIMARY,
                     font=("Segoe UI", 16, "bold")).pack(anchor="w")
            tk.Label(inner, text=title, bg=BG_SURFACE, fg=TEXT_MUTED,
                     font=("Segoe UI", 9)).pack(anchor="w")

    def _build_step_indicator(self):
        for w in self._step_frame.winfo_children():
            w.destroy()
        has_upload = len(self.uploaded_paths) > 0
        has_process = len(self.formations_data) > 0
        has_excel = self.excel_bytes is not None
        steps = [
            ("1", "Upload Gambar", has_upload),
            ("2", "Proses", has_process),
            ("3", "Export Excel", has_excel),
        ]
        for i, (num, label, done) in enumerate(steps):
            sf = tk.Frame(self._step_frame, bg=BG_BASE)
            sf.pack(side="left", padx=(0, 4))
            color = CLR_SUCCESS if done else BORDER
            cv = tk.Canvas(sf, width=26, height=26, bg=BG_BASE, highlightthickness=0)
            cv.create_oval(1, 1, 25, 25, fill=color, outline="")
            cv.create_text(13, 13, text=num, fill="white" if done else TEXT_MUTED,
                           font=("Segoe UI", 9, "bold"))
            cv.pack(side="left", padx=(0, 4))
            fg = TEXT_PRIMARY if done else TEXT_MUTED
            tk.Label(sf, text=label, bg=BG_BASE, fg=fg, font=("Segoe UI", 10)).pack(side="left")
            if i < 2:
                tk.Label(self._step_frame, text="->", bg=BG_BASE, fg=TEXT_MUTED,
                         font=("Segoe UI", 10)).pack(side="left", padx=6)

    def _build_tab_row(self):
        for w in self._tab_row.winfo_children():
            w.destroy()
        self._tab_labels = {}
        self._tab_indicators = {}
        for i in range(1, N_FORMATIONS + 1):
            tf = tk.Frame(self._tab_row, bg=BG_BASE)
            tf.pack(side="left", padx=1)
            lbl = tk.Label(tf, text=f"F{i}", bg=BG_BASE, fg=TEXT_MUTED,
                           font=("Segoe UI", 10), padx=10, pady=4, cursor="hand2")
            lbl.pack()
            lbl.bind("<Button-1>", lambda e, idx=i: self._select_tab(idx))
            ind = tk.Frame(tf, bg=BG_BASE, height=3)
            ind.pack(fill="x")
            self._tab_labels[i] = lbl
            self._tab_indicators[i] = ind

    def _select_tab(self, idx):
        if idx not in self.formations_data:
            return
        self._active_tab = idx
        for i, lbl in self._tab_labels.items():
            active = (i == idx)
            lbl.config(fg=PRIMARY if active else TEXT_MUTED,
                       font=("Segoe UI", 10, "bold") if active else ("Segoe UI", 10))
            self._tab_indicators[i].config(bg=PRIMARY if active else BG_BASE)
        self._show_formation(idx)

    def _show_empty_state(self):
        for w in self._formation_frame.winfo_children():
            w.destroy()
        tk.Label(self._formation_frame,
                 text="Mulai dengan mengunggah gambar formasi di panel kiri,\nlalu klik Proses Semua Gambar.",
                 bg=BG_BASE, fg=TEXT_MUTED, font=("Segoe UI", 11), justify="center").pack(
                     expand=True, pady=60)


    def _show_formation(self, fi):
        for w in self._formation_frame.winfo_children():
            w.destroy()
        if fi not in self.formations_data:
            self._show_empty_state()
            return
        fdata = self.formations_data[fi]

        self._formation_frame.columnconfigure(0, weight=0, minsize=260)
        self._formation_frame.columnconfigure(1, weight=1)
        self._formation_frame.rowconfigure(0, weight=1)


        left = tk.Frame(self._formation_frame, bg=BG_SURFACE, highlightbackground=BORDER,
                        highlightthickness=1, width=260)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 8), pady=4)
        left.grid_propagate(False)

        tk.Label(left, text="Gambar Asli", bg=BG_SURFACE, fg=TEXT_PRIMARY,
                 font=("Segoe UI", 11, "bold")).pack(padx=12, pady=(10, 4), anchor="w")
        thumb_frame = tk.Frame(left, bg=BG_SURFACE, highlightbackground=BORDER, highlightthickness=1)
        thumb_frame.pack(padx=12, pady=4)
        try:
            orig_path = self.uploaded_paths.get(fi)
            if orig_path:
                img = Image.open(orig_path).convert("RGB")
                img.thumbnail((220, 180), Image.LANCZOS)
                photo = ImageTk.PhotoImage(img)
                lbl = tk.Label(thumb_frame, image=photo, bg=BG_SURFACE)
                lbl.pack(padx=2, pady=2)
                self._thumb_refs[fi] = photo
        except Exception:
            tk.Label(thumb_frame, text="[Tidak tersedia]", bg=BG_SURFACE, fg=TEXT_MUTED,
                     font=("Segoe UI", 9)).pack(padx=8, pady=8)


        tk.Label(left, text="Distribusi Warna", bg=BG_SURFACE, fg=TEXT_PRIMARY,
                 font=("Segoe UI", 11, "bold")).pack(padx=12, pady=(12, 4), anchor="w")
        total = sum(fdata["distribution"].values())
        max_cnt = max(fdata["distribution"].values()) if fdata["distribution"] else 1
        for wkey in WARNA_KEYS:
            cnt = fdata["distribution"].get(wkey, 0)
            pct = cnt / total * 100 if total > 0 else 0
            info = WARNA_DEFINISI[wkey]
            rf = tk.Frame(left, bg=BG_SURFACE)
            rf.pack(fill="x", padx=12, pady=2)
            tk.Label(rf, text=info["label"], bg=BG_SURFACE, fg=TEXT_PRIMARY,
                     font=("Segoe UI", 9), width=8, anchor="w").pack(side="left")
            bar_w = 120
            bar_cv = tk.Canvas(rf, width=bar_w, height=14, bg=BORDER, highlightthickness=0)
            bar_cv.pack(side="left", padx=4)
            fill_w = int(cnt / max_cnt * bar_w) if max_cnt > 0 else 0
            bar_cv.create_rectangle(0, 0, fill_w, 14, fill=info["hex"], outline="")
            tk.Label(rf, text=f"{cnt} ({pct:.1f}%)", bg=BG_SURFACE, fg=TEXT_MUTED,
                     font=("Segoe UI", 8)).pack(side="left")


        right = tk.Frame(self._formation_frame, bg=BG_BASE)
        right.grid(row=0, column=1, sticky="nsew", padx=(8, 0), pady=4)
        right.columnconfigure(0, weight=1)
        right.rowconfigure(0, weight=1)

        fig = render_grid_preview(fdata, f"Formasi {fi}",
                                  show_grid_lines=self.show_grid_var.get())
        canvas = FigureCanvasTkAgg(fig, master=right)
        canvas.draw()
        widget = canvas.get_tk_widget()
        widget.config(bg=BG_BASE, highlightthickness=0)
        widget.grid(row=0, column=0, sticky="nsew")
        self._canvas_refs[fi] = canvas
        plt.close(fig)


    def _pick_image(self, idx):

        if idx in self.formations_data:
            proceed = messagebox.askyesno("Konfirmasi",
                "Formasi ini sudah diproses. Upload gambar baru "
                "akan menghapus hasil sebelumnya. Lanjut?")
            if not proceed:
                return

        path = filedialog.askopenfilename(
            title=f"Pilih gambar untuk Formasi {idx}",
            filetypes=[("Gambar", "*.jpg *.jpeg *.png *.webp"), ("Semua file", "*.*")])
        if not path:
            return


        import os
        ext = os.path.splitext(path)[1].lower()
        if ext not in ALLOWED_EXT:
            self._slot_err_labels[idx].config(text=f"Format {ext} tidak didukung. Gunakan JPG/PNG/WEBP.")
            self._slot_err_labels[idx].pack(fill="x", pady=(1, 0))
            return


        self._slot_err_labels[idx].config(text="")
        self._slot_err_labels[idx].pack_forget()


        if idx in self.formations_data:
            del self.formations_data[idx]

        self.uploaded_paths[idx] = path
        fname = path.split("/")[-1].split("\\")[-1]
        if len(fname) > 20:
            fname = fname[:17] + "..."
        self._slot_fname_labels[idx].config(text=fname, fg=TEXT_INVERSE)
        self._refresh_sidebar_state()

    def _remove_formation(self, idx):
        if idx in self.uploaded_paths:
            del self.uploaded_paths[idx]
        if idx in self.formations_data:
            del self.formations_data[idx]
        self._slot_fname_labels[idx].config(text="Belum diunggah", fg="#A8A29E")
        self._slot_err_labels[idx].config(text="")
        self._slot_err_labels[idx].pack_forget()
        self._refresh_sidebar_state()
        self._build_metrics()
        self._build_step_indicator()
        self._build_tab_row()
        if self._active_tab == idx:
            self._active_tab = None
            self._show_empty_state()

    def _refresh_sidebar_state(self):
        n = len(self.uploaded_paths)
        self._upload_count_lbl.config(text=f"{n} / {N_FORMATIONS} diunggah")

        for i in range(1, N_FORMATIONS + 1):
            dot = self._slot_dots[i]
            dot.delete("dot")
            if i in self.formations_data:
                color = CLR_SUCCESS
            elif i in self.uploaded_paths:
                color = CLR_WARNING
            else:
                color = BORDER
            dot.create_oval(2, 2, 10, 10, fill=color, outline="", tags="dot")


        if n > 0:
            self._process_btn.config(bg=PRIMARY, cursor="hand2")
            self._process_tip.config(text="")
        else:
            self._process_btn.config(bg="#A8A29E", cursor="arrow")
            self._process_tip.config(text="Upload minimal 1 gambar untuk memulai.")

        self._build_step_indicator()


    def _start_processing(self):
        if not self.uploaded_paths:
            return

        self._proc_frame.pack(fill="x", padx=24, pady=4,
                              before=self._formation_frame.master.winfo_children()[3]
                              if len(self._content.winfo_children()) > 3 else None)
        self._proc_frame.pack(fill="x", padx=24, pady=4)
        self._proc_bar_canvas.delete("fill")
        self._proc_label.config(text="Memulai pemrosesan...")
        self._process_btn.config(bg="#A8A29E", cursor="arrow")
        self.update_idletasks()

        thread = threading.Thread(target=self._process_thread, daemon=True)
        thread.start()

    def _process_thread(self):
        self.formations_data = {}
        errors = []
        sorted_keys = sorted(self.uploaded_paths.keys())
        for step, fi in enumerate(sorted_keys, start=1):
            self.after(0, lambda s=step, i=fi, t=len(sorted_keys): self._update_progress(s, i, t))
            try:
                fdata = process_formation_image(self.uploaded_paths[fi])
                self.formations_data[fi] = fdata
            except Exception as e:
                errors.append(f"Formasi {fi}: {e}")
        self.after(0, lambda: self._processing_done(errors))

    def _update_progress(self, step, fi, total):
        self._proc_label.config(text=f"Memproses Formasi {fi} ({step}/{total})...")
        w = self._proc_bar_canvas.winfo_width()
        fill_w = int(step / total * w) if total > 0 else 0
        self._proc_bar_canvas.delete("fill")
        self._proc_bar_canvas.create_rectangle(0, 0, fill_w, 6, fill=PRIMARY, outline="", tags="fill")
        self.update_idletasks()

    def _processing_done(self, errors):
        self._proc_label.config(text="Semua formasi selesai diproses.")
        w = self._proc_bar_canvas.winfo_width()
        self._proc_bar_canvas.delete("fill")
        self._proc_bar_canvas.create_rectangle(0, 0, w, 6, fill=CLR_SUCCESS, outline="", tags="fill")
        self._process_btn.config(bg=PRIMARY, cursor="hand2")

        if errors:
            messagebox.showerror("Error Pemrosesan",
                "Terjadi error pada beberapa formasi:\n\n" + "\n".join(errors))
        if not self.formations_data:
            messagebox.showerror("Error", "Tidak ada formasi yang berhasil diproses.")
            return

        self._build_metrics()
        self._build_step_indicator()
        self._refresh_sidebar_state()
        self._build_tab_row()
        self._rebuild_dist_table()
        self.excel_bytes = None
        self._dl_btn.config(bg="#D6D3D1")
        self._dl_btn_enabled = False
        self._excel_status.config(text="")


        last = max(self.formations_data.keys())
        self._select_tab(last)


    def _rebuild_dist_table(self):
        for w in self._dist_frame.winfo_children():
            w.destroy()
        self._dist_tree = None
        if not self.formations_data:
            return
        cols = ["Warna"] + [f"F{fi}" for fi in sorted(self.formations_data.keys())] + ["Total"]
        tree = ttk.Treeview(self._dist_frame, columns=cols, show="headings",
                             height=len(WARNA_KEYS))
        for col in cols:
            w = 90 if col == "Warna" else 60
            tree.heading(col, text=col)
            tree.column(col, width=w, anchor="center", stretch=False)
        tree.tag_configure("alt", background=BG_BASE)
        for wi, wkey in enumerate(WARNA_KEYS):
            lbl = WARNA_DEFINISI[wkey]["label"]
            total = 0
            row = [lbl]
            for fi in sorted(self.formations_data.keys()):
                cnt = self.formations_data[fi]["distribution"].get(wkey, 0)
                total += cnt
                row.append(cnt)
            row.append(total)
            tag = ("alt",) if wi % 2 == 1 else ()
            tree.insert("", "end", values=row, tags=tag)
        xsb = ttk.Scrollbar(self._dist_frame, orient="horizontal", command=tree.xview)
        tree.configure(xscrollcommand=xsb.set)
        tree.pack(fill="x", side="top")
        xsb.pack(fill="x", side="top")
        self._dist_tree = tree


    def _generate_excel(self):
        if not self.formations_data:
            messagebox.showwarning("Peringatan",
                "Belum ada formasi yang diproses.\nProses gambar terlebih dahulu.")
            return
        n = len(self.formations_data)
        if n < N_FORMATIONS:
            proceed = messagebox.askyesno("Konfirmasi",
                f"Baru {n} dari {N_FORMATIONS} formasi diproses.\n"
                "Tetap generate Excel dengan data yang ada?")
            if not proceed:
                return
        self._excel_status.config(text="Sedang memproses...", fg=TEXT_MUTED)
        self._gen_btn.config(fg="#A8A29E")
        self.update_idletasks()
        thread = threading.Thread(target=self._excel_thread, daemon=True)
        thread.start()

    def _excel_thread(self):
        try:
            bg_label = self._bg_combo_var.get()
            bg_key = self._bg_display_map.get(bg_label, "putih")
            ordered = [self.formations_data[i] for i in sorted(self.formations_data.keys())]
            df = build_participant_dataframe(ordered, bg_color=bg_key)
            xls_bytes = export_to_excel(df, ordered)
            self.excel_bytes = xls_bytes
            self._df_preview = df
            self.after(0, self._excel_done)
        except Exception as e:
            self.after(0, lambda err=str(e): self._excel_error(err))

    def _excel_done(self):
        n = len(self.formations_data)
        total_entries = GRID_SIZE * GRID_SIZE * n
        self._excel_status.config(
            text=f"Siap diunduh. {GRID_SIZE*GRID_SIZE:,} peserta x {n} formasi = {total_entries:,} entri.",
            fg=CLR_SUCCESS)
        self._dl_btn.config(bg=PRIMARY, cursor="hand2")
        self._dl_btn_enabled = True
        self._dl_btn.bind("<Button-1>", lambda e: self._save_excel())
        self._gen_btn.config(fg=PRIMARY)
        self._build_step_indicator()
        self._rebuild_preview_table()

    def _excel_error(self, err):
        messagebox.showerror("Error Excel", f"Gagal membuat file Excel:\n{err}")
        self._excel_status.config(text="Gagal.", fg=CLR_ERROR)
        self._gen_btn.config(fg=PRIMARY)

    def _save_excel(self):
        if not self.excel_bytes:
            return
        path = filedialog.asksaveasfilename(
            title="Simpan File Excel", defaultextension=".xlsx",
            initialfile="ODM_Undip_2026_Mozaik_Formation.xlsx",
            filetypes=[("Excel Workbook", "*.xlsx"), ("Semua file", "*.*")])
        if not path:
            return
        try:
            with open(path, "wb") as f:
                f.write(self.excel_bytes)
            messagebox.showinfo("Berhasil", f"File Excel berhasil disimpan:\n{path}")
        except Exception as e:
            messagebox.showerror("Error", f"Gagal menyimpan file:\n{e}")


    def _rebuild_preview_table(self):
        for w in self._preview_frame.winfo_children():
            w.destroy()
        self._preview_tree = None
        if not hasattr(self, "_df_preview") or self._df_preview is None:
            return
        df = self._df_preview.head(10)
        cols = list(df.columns)
        frame = tk.Frame(self._preview_frame, bg=BG_BASE)
        frame.pack(fill="x")
        xscroll = ttk.Scrollbar(frame, orient="horizontal")
        yscroll = ttk.Scrollbar(frame, orient="vertical")
        tree = ttk.Treeview(frame, columns=cols, show="headings", height=10,
                             xscrollcommand=xscroll.set, yscrollcommand=yscroll.set)
        xscroll.config(command=tree.xview)
        yscroll.config(command=tree.yview)
        tree.tag_configure("alt", background=BG_BASE)
        for col in cols:
            w = 100 if "Komando" in col else (80 if col.endswith("_Warna") or col.endswith("_Posisi") else 70)
            tree.heading(col, text=col)
            tree.column(col, width=w, anchor="center", stretch=False)
        for ri, (_, row) in enumerate(df.iterrows()):
            tag = ("alt",) if ri % 2 == 1 else ()
            tree.insert("", "end", values=list(row), tags=tag)
        tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(0, weight=1)
        self._preview_tree = tree


def main():
    app = ODMApp()
    app.mainloop()


if __name__ == "__main__":
    main()
